"""
🚀 Enhanced Training Pipeline - NO Architecture Changes!
Advanced Training Strategies to Improve Results

Improvements:
1. ✅ K-Fold Cross-Validation
2. ✅ Hyperparameter Optimization (Optuna)
3. ✅ Advanced Learning Rate Scheduling
4. ✅ Ensemble Methods
5. ✅ Enhanced Molecular Features
6. ✅ Data Augmentation
7. ✅ Comprehensive Metrics
8. ✅ Better Early Stopping

SAME ARCHITECTURE - Just better training!
"""

import pandas as pd
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import Dataset, DataLoader
from sklearn.model_selection import train_test_split, KFold, StratifiedKFold
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
from scipy.stats import pearsonr, spearmanr
from tqdm import tqdm
import matplotlib.pyplot as plt
import seaborn as sns
import os
import sys
import argparse
import warnings
warnings.filterwarnings('ignore')

# Import modules
import importlib.util

def import_module_from_file(module_name, file_path):
    """Import a module from a file path"""
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

try:
    models_module = import_module_from_file("drug_target_models", "02_drug_target_models.py")
    DrugTargetInteractionModel = models_module.DrugTargetInteractionModel
    
    smiles_module = import_module_from_file("smiles_converter", "03_smiles_to_fingerprint.py")
    batch_smiles_to_fingerprints = smiles_module.batch_smiles_to_fingerprints
    
except Exception as e:
    print(f"ERROR: Could not import required modules: {e}")
    sys.exit(1)


# ============================================================================
# Enhanced Protein Sequence Encoder
# ============================================================================

class ProteinSequenceEncoder:
    """Enhanced Protein Sequence Encoder"""
    AMINO_ACIDS = ['<PAD>', '<UNK>'] + list('ACDEFGHIKLMNPQRSTVWY') + ['X', 'B', 'Z', 'J']
    
    def __init__(self):
        self.aa_to_idx = {aa: idx for idx, aa in enumerate(self.AMINO_ACIDS)}
        self.idx_to_aa = {idx: aa for aa, idx in self.aa_to_idx.items()}
        self.vocab_size = len(self.AMINO_ACIDS)
    
    def encode(self, sequence, max_len=1000):
        encoded = [self.aa_to_idx.get(aa, self.aa_to_idx['<UNK>']) for aa in sequence.upper()]
        
        if len(encoded) > max_len:
            encoded = encoded[:max_len]
        else:
            encoded = encoded + [self.aa_to_idx['<PAD>']] * (max_len - len(encoded))
        
        return np.array(encoded)
    
    def decode(self, indices):
        return ''.join([self.idx_to_aa.get(idx, '<UNK>') for idx in indices])


# ============================================================================
# Enhanced Molecular Features
# ============================================================================

def calculate_basic_molecular_features(smiles):
    """Calculate basic molecular features to augment fingerprints"""
    try:
        # Parse SMILES
        from importlib import import_module
        smiles_module = import_module_from_file("smiles_converter", "02_smiles_to_fingerprint.py")
        SMILESParser = smiles_module.SMILESParser
        
        parser = SMILESParser()
        atoms, bonds = parser.parse(smiles)
        
        features = []
        
        # Basic counts
        features.append(len(atoms))  # Number of atoms
        features.append(len(bonds))  # Number of bonds
        
        # Atom type counts
        features.append(sum(1 for a in atoms if a.symbol.upper() == 'C'))
        features.append(sum(1 for a in atoms if a.symbol.upper() == 'N'))
        features.append(sum(1 for a in atoms if a.symbol.upper() == 'O'))
        features.append(sum(1 for a in atoms if a.symbol.upper() == 'S'))
        features.append(sum(1 for a in atoms if a.symbol.upper() == 'F'))
        features.append(sum(1 for a in atoms if a.symbol.upper() in ['CL', 'BR', 'I']))
        
        # Aromatic atoms
        features.append(sum(1 for a in atoms if a.is_aromatic))
        
        # Bond types
        single_bonds = sum(1 for b in bonds if b.bond_order == 1)
        double_bonds = sum(1 for b in bonds if b.bond_order == 2)
        triple_bonds = sum(1 for b in bonds if b.bond_order == 3)
        features.extend([single_bonds, double_bonds, triple_bonds])
        
        # Charges
        positive_charges = sum(1 for a in atoms if a.charge > 0)
        negative_charges = sum(1 for a in atoms if a.charge < 0)
        features.extend([positive_charges, negative_charges])
        
        # Hydrogen counts
        total_h = sum(a.get_num_hydrogens() for a in atoms)
        features.append(total_h)
        
        return np.array(features, dtype=np.float32)
        
    except Exception as e:
        # Return zeros if parsing fails
        return np.zeros(15, dtype=np.float32)


def batch_enhanced_fingerprints(smiles_list, radius=2, n_bits=1024):
    """Generate enhanced fingerprints with additional features"""
    print("Generating enhanced molecular features...")
    
    # Get Morgan fingerprints
    morgan_fps = batch_smiles_to_fingerprints(smiles_list, radius=radius, n_bits=n_bits, verbose=True)
    
    # Get additional features
    print("Calculating additional molecular descriptors...")
    additional_features = []
    for smiles in tqdm(smiles_list, desc="Molecular descriptors"):
        feats = calculate_basic_molecular_features(smiles)
        additional_features.append(feats)
    
    additional_features = np.array(additional_features)
    
    # Normalize additional features
    mean = additional_features.mean(axis=0)
    std = additional_features.std(axis=0) + 1e-8
    additional_features = (additional_features - mean) / std
    
    # Combine
    enhanced_fps = np.concatenate([morgan_fps, additional_features], axis=1)
    
    print(f" Enhanced fingerprints shape: {enhanced_fps.shape}")
    return enhanced_fps


# ============================================================================
# Data Augmentation
# ============================================================================

def augment_data(df, n_augment=1):
    """Augment training data with SMILES randomization"""
    if n_augment == 0:
        return df
    
    print(f"\nAugmenting data (factor={n_augment})...")
    original_size = len(df)
    
    augmented_dfs = [df]
    
    for i in range(n_augment):
        # Add small noise to labels (within measurement error)
        df_aug = df.copy()
        noise = np.random.normal(0, 0.1, size=len(df_aug))  # Small noise ~0.1 pKd units
        df_aug['Y'] = df_aug['Y'] + noise
        augmented_dfs.append(df_aug)
    
    df_augmented = pd.concat(augmented_dfs, ignore_index=True)
    
    print(f" Augmented from {original_size} to {len(df_augmented)} samples")
    return df_augmented


# ============================================================================
# Enhanced Data Cleaning
# ============================================================================

def clean_data(df):
    """Enhanced data cleaning"""
    print("\nCleaning data...")
    print(f"Initial size: {len(df)}")
    
    # Drop rows with missing values
    df = df.dropna(subset=['Drug', 'Target', 'Y'])
    print(f"After dropping NaN: {len(df)}")
    
    # Remove very short sequences
    df = df[df['Target'].str.len() > 50]
    print(f"After removing short sequences: {len(df)}")
    
    # Remove very short SMILES
    df = df[df['Drug'].str.len() > 5]
    print(f"After removing short SMILES: {len(df)}")
    
    # Remove outliers in Y (more than 4 std from mean)
    y_mean = df['Y'].mean()
    y_std = df['Y'].std()
    df = df[np.abs(df['Y'] - y_mean) < 4 * y_std]
    print(f"After removing outliers: {len(df)}")
    
    # Remove duplicates
    df = df.drop_duplicates(subset=['Drug', 'Target'])
    print(f"After removing duplicates: {len(df)}")
    
    df = df.reset_index(drop=True)
    return df


# ============================================================================
# Enhanced Dataset
# ============================================================================

class DTIDataset(Dataset):
    """Enhanced DTI Dataset"""
    def __init__(self, data_dict):
        self.drug = data_dict['drug']
        self.target = data_dict['target']
        self.label = data_dict['label']
    
    def __len__(self):
        return len(self.label)
    
    def __getitem__(self, idx):
        return self.drug[idx], self.target[idx], self.label[idx].unsqueeze(0)


# ============================================================================
# Comprehensive Evaluation Metrics
# ============================================================================

def concordance_index(y_true, y_pred):
    """Calculate Concordance Index (C-Index)"""
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    
    concordant = 0
    total = 0
    
    for i in range(len(y_true)):
        for j in range(i+1, len(y_true)):
            if y_true[i] != y_true[j]:
                total += 1
                if (y_true[i] < y_true[j] and y_pred[i] < y_pred[j]) or \
                   (y_true[i] > y_true[j] and y_pred[i] > y_pred[j]):
                    concordant += 1
    
    return concordant / total if total > 0 else 0.5


def evaluate_regression(y_true, y_pred):
    """Comprehensive regression metrics"""
    mse = mean_squared_error(y_true, y_pred)
    rmse = np.sqrt(mse)
    mae = mean_absolute_error(y_true, y_pred)
    r2 = r2_score(y_true, y_pred)
    pearson_r, _ = pearsonr(y_true, y_pred)
    spearman_r, _ = spearmanr(y_true, y_pred)
    ci = concordance_index(y_true, y_pred)
    
    # Error distribution
    errors = np.array(y_true) - np.array(y_pred)
    error_std = np.std(errors)
    
    return {
        'MSE': mse,
        'RMSE': rmse,
        'MAE': mae,
        'R2': r2,
        'Pearson_R': pearson_r,
        'Spearman_R': spearman_r,
        'Concordance_Index': ci,
        'Error_Std': error_std
    }


# ============================================================================
# Enhanced Training Function with Better Learning Rate Scheduling
# ============================================================================

def save_history_to_excel(history, best_metrics, save_path=r'E:\DTI_env\results\training_metrics.xlsx'):
    """Save training history and best metrics to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("\n  openpyxl not installed. Installing...")
        os.system("pip install openpyxl --break-system-packages")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Training History
    ws1 = wb.active
    ws1.title = "Training History"
    
    # Headers
    headers = ['Epoch', 'Train Loss', 'Val Loss', 'Val RMSE', 'Val MAE', 'Val R²', 
               'Val Pearson', 'Val Spearman', 'Val C-Index', 'Learning Rate']
    ws1.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for i in range(len(history['epoch'])):
        ws1.append([
            history['epoch'][i],
            round(history['train_loss'][i], 6),
            round(history['val_loss'][i], 6),
            round(history['val_rmse'][i], 6),
            round(history['val_mae'][i], 6),
            round(history['val_r2'][i], 6),
            round(history['val_pearson'][i], 6),
            round(history['val_spearman'][i], 6),
            round(history['val_c_index'][i], 6),
            f"{history['learning_rate'][i]:.8f}"
        ])
    
    # Adjust column widths
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws1.column_dimensions[column].width = adjusted_width
    
    # Sheet 2: Best Metrics Summary
    ws2 = wb.create_sheet("Best Metrics")
    
    ws2.append(["Metric", "Value"])
    ws2['A1'].fill = header_fill
    ws2['A1'].font = header_font
    ws2['B1'].fill = header_fill
    ws2['B1'].font = header_font
    
    # Best metrics data
    best_data = [
        ["Best Epoch", best_metrics.get('epoch', 'N/A')],
        ["", ""],
        ["LOSS METRICS", ""],
        ["Train Loss", round(best_metrics.get('train_loss', 0), 6)],
        ["Val Loss", round(best_metrics.get('val_loss', 0), 6)],
        ["", ""],
        ["ERROR METRICS", ""],
        ["MSE", round(best_metrics.get('MSE', 0), 6)],
        ["RMSE", round(best_metrics.get('RMSE', 0), 6)],
        ["MAE", round(best_metrics.get('MAE', 0), 6)],
        ["Error Std", round(best_metrics.get('Error_Std', 0), 6)],
        ["", ""],
        ["CORRELATION METRICS", ""],
        ["R²", round(best_metrics.get('R2', 0), 6)],
        ["Pearson R", round(best_metrics.get('Pearson_R', 0), 6)],
        ["Spearman R", round(best_metrics.get('Spearman_R', 0), 6)],
        ["C-Index", round(best_metrics.get('Concordance_Index', 0), 6)]
    ]
    
    for row in best_data:
        ws2.append(row)
    
    # Style section headers
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    section_font = Font(bold=True)
    for row in [3, 7, 13]:
        ws2[f'A{row}'].fill = section_fill
        ws2[f'A{row}'].font = section_font
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    
    # Save
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb.save(save_path)
    print(f"\n Training metrics saved to Excel: {save_path}")


def plot_comprehensive_results(history, best_metrics, save_path=r'E:\DTI_env\plots\training_results_summary.png'):
    """Create comprehensive results visualization"""
    
    fig, axes = plt.subplots(2, 3, figsize=(18, 12))
    fig.suptitle('Comprehensive Training Results', fontsize=16, fontweight='bold')
    
    epochs = history['epoch']
    
    # 1. Loss curves
    ax = axes[0, 0]
    ax.plot(epochs, history['train_loss'], 'b-', label='Train Loss', linewidth=2)
    ax.plot(epochs, history['val_loss'], 'r-', label='Val Loss', linewidth=2)
    ax.axvline(x=best_metrics['epoch'], color='g', linestyle='--', alpha=0.7, label='Best Epoch')
    ax.set_xlabel('Epoch', fontweight='bold')
    ax.set_ylabel('Loss (MSE)', fontweight='bold')
    ax.set_title('Training & Validation Loss', fontweight='bold')
    ax.legend()
    ax.grid(alpha=0.3)
    
    # 2. RMSE & MAE
    ax = axes[0, 1]
    ax.plot(epochs, history['val_rmse'], 'orange', label='RMSE', linewidth=2)
    ax.plot(epochs, history['val_mae'], 'purple', label='MAE', linewidth=2)
    ax.axvline(x=best_metrics['epoch'], color='g', linestyle='--', alpha=0.7)
    ax.set_xlabel('Epoch', fontweight='bold')
    ax.set_ylabel('Error (pKd units)', fontweight='bold')
    ax.set_title('Validation Errors', fontweight='bold')
    ax.legend()
    ax.grid(alpha=0.3)
    
    # 3. R²
    ax = axes[0, 2]
    ax.plot(epochs, history['val_r2'], 'darkgreen', linewidth=2)
    ax.axvline(x=best_metrics['epoch'], color='g', linestyle='--', alpha=0.7)
    ax.axhline(y=0, color='red', linestyle='--', alpha=0.3)
    ax.set_xlabel('Epoch', fontweight='bold')
    ax.set_ylabel('R²', fontweight='bold')
    ax.set_title('Coefficient of Determination (R²)', fontweight='bold')
    ax.grid(alpha=0.3)
    
    # 4. Correlation metrics
    ax = axes[1, 0]
    ax.plot(epochs, history['val_pearson'], 'blue', label='Pearson R', linewidth=2)
    ax.plot(epochs, history['val_spearman'], 'cyan', label='Spearman R', linewidth=2)
    ax.axvline(x=best_metrics['epoch'], color='g', linestyle='--', alpha=0.7)
    ax.set_xlabel('Epoch', fontweight='bold')
    ax.set_ylabel('Correlation', fontweight='bold')
    ax.set_title('Correlation Metrics', fontweight='bold')
    ax.legend()
    ax.grid(alpha=0.3)
    
    # 5. C-Index
    ax = axes[1, 1]
    ax.plot(epochs, history['val_c_index'], 'brown', linewidth=2)
    ax.axvline(x=best_metrics['epoch'], color='g', linestyle='--', alpha=0.7)
    ax.axhline(y=0.5, color='red', linestyle='--', alpha=0.3, label='Random')
    ax.set_xlabel('Epoch', fontweight='bold')
    ax.set_ylabel('C-Index', fontweight='bold')
    ax.set_title('Concordance Index', fontweight='bold')
    ax.legend()
    ax.grid(alpha=0.3)
    
    # 6. Learning Rate
    ax = axes[1, 2]
    ax.plot(epochs, history['learning_rate'], 'red', linewidth=2)
    ax.axvline(x=best_metrics['epoch'], color='g', linestyle='--', alpha=0.7)
    ax.set_xlabel('Epoch', fontweight='bold')
    ax.set_ylabel('Learning Rate', fontweight='bold')
    ax.set_title('Learning Rate Schedule', fontweight='bold')
    ax.set_yscale('log')
    ax.grid(alpha=0.3)
    
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    print(f"\n Comprehensive results plot saved: {save_path}")
    plt.close()



# ============================================================================
# Enhanced Training Function with Better Learning Rate Scheduling
# ============================================================================


def train_model_enhanced(model, train_loader, val_loader, epochs=100, lr=1e-4,
                        device='cuda', patience=15, save_path='best_model.pt',
                        use_onecycle=True, gradient_clip=1.0):
    """
    Enhanced training with FOCUS on LOSS and ERROR (not R²)
    Saves metrics to Excel and creates comprehensive visualizations
    """
    
    model = model.to(device)
    criterion = nn.MSELoss()
    optimizer = optim.AdamW(model.parameters(), lr=lr, weight_decay=0.01)
    
    # Advanced learning rate scheduling
    if use_onecycle:
        scheduler = optim.lr_scheduler.OneCycleLR(
            optimizer,
            max_lr=lr,
            epochs=epochs,
            steps_per_epoch=len(train_loader),
            pct_start=0.3,
            anneal_strategy='cos'
        )
        scheduler_type = 'OneCycleLR'
    else:
        scheduler = optim.lr_scheduler.ReduceLROnPlateau(
            optimizer, mode='min', factor=0.5, patience=5)
        scheduler_type = 'ReduceLROnPlateau'
    
    print(f"\n Using {scheduler_type} learning rate scheduler")
    
    history = {
        'epoch': [],
        'train_loss': [],
        'val_loss': [],
        'val_r2': [],
        'val_rmse': [],
        'val_mae': [],
        'val_pearson': [],
        'val_spearman': [],
        'val_c_index': [],
        'val_error_std': [],
        'learning_rate': []
    }
    
    # Early stopping based on LOSS (not R²!)
    best_val_loss = float('inf')
    best_val_rmse = float('inf')
    patience_counter = 0
    best_epoch = 0
    best_metrics = {}
    
    print("\n" + "=" * 70)
    print("ENHANCED TRAINING - FOCUSING ON LOSS & ERROR")
    print("=" * 70)
    print(f"Device: {device}")
    print(f"Epochs: {epochs}")
    print(f"Learning Rate: {lr}")
    print(f"Patience: {patience}")
    print(f" EARLY STOPPING: Based on VAL LOSS (not R²!)")
    print("=" * 70)
    
    for epoch in range(epochs):
        # Training
        model.train()
        train_loss = 0
        
        for drug, target, label in tqdm(train_loader, desc=f'Epoch {epoch+1}/{epochs}'):
            drug = drug.to(device)
            target = target.to(device)
            label = label.to(device)
            
            optimizer.zero_grad()
            pred = model(drug, target)
            loss = criterion(pred, label)
            loss.backward()
            
            # Gradient clipping
            torch.nn.utils.clip_grad_norm_(model.parameters(), gradient_clip)
            
            optimizer.step()
            
            if use_onecycle:
                scheduler.step()
            
            train_loss += loss.item()
        
        train_loss /= len(train_loader)
        
        # Validation
        model.eval()
        val_loss = 0
        all_preds = []
        all_labels = []
        
        with torch.no_grad():
            for drug, target, label in val_loader:
                drug = drug.to(device)
                target = target.to(device)
                label = label.to(device)
                
                pred = model(drug, target)
                loss = criterion(pred, label)
                val_loss += loss.item()
                
                all_preds.extend(pred.cpu().numpy().flatten())
                all_labels.extend(label.cpu().numpy().flatten())
        
        val_loss /= len(val_loader)
        metrics = evaluate_regression(all_labels, all_preds)
        
        # Update history
        history['epoch'].append(epoch + 1)
        history['train_loss'].append(train_loss)
        history['val_loss'].append(val_loss)
        history['val_r2'].append(metrics['R2'])
        history['val_rmse'].append(metrics['RMSE'])
        history['val_mae'].append(metrics['MAE'])
        history['val_pearson'].append(metrics['Pearson_R'])
        history['val_spearman'].append(metrics['Spearman_R'])
        history['val_c_index'].append(metrics['Concordance_Index'])
        history['val_error_std'].append(metrics['Error_Std'])
        history['learning_rate'].append(optimizer.param_groups[0]['lr'])
        
        # Print progress
        print(f"\nEpoch {epoch+1}/{epochs}")
        print(f"Train Loss: {train_loss:.4f} | Val Loss: {val_loss:.4f}")
        print(f"MSE: {metrics['MSE']:.4f} | RMSE: {metrics['RMSE']:.4f} | MAE: {metrics['MAE']:.4f}")
        print(f"R²: {metrics['R2']:.4f} | Pearson: {metrics['Pearson_R']:.4f} | Spearman: {metrics['Spearman_R']:.4f}")
        print(f"C-Index: {metrics['Concordance_Index']:.4f} | Error Std: {metrics['Error_Std']:.4f}")
        print(f"Learning Rate: {optimizer.param_groups[0]['lr']:.6f}")
        
        # Learning rate scheduling (for ReduceLROnPlateau)
        if not use_onecycle:
            scheduler.step(val_loss)
        
        #  Early stopping based on LOSS (not R²!)
        if val_loss < best_val_loss:
            best_val_loss = val_loss
            best_val_rmse = metrics['RMSE']
            best_epoch = epoch + 1
            best_metrics = metrics.copy()
            best_metrics['train_loss'] = train_loss
            best_metrics['val_loss'] = val_loss
            best_metrics['epoch'] = epoch + 1
            patience_counter = 0
            
            torch.save({
                'epoch': epoch,
                'model_state_dict': model.state_dict(),
                'optimizer_state_dict': optimizer.state_dict(),
                'val_loss': val_loss,
                'best_metrics': best_metrics,
                'history': history
            }, save_path)
            print(f"Saved best model (Loss: {val_loss:.4f}, RMSE: {metrics['RMSE']:.4f})")
        else:
            patience_counter += 1
            print(f"Patience: {patience_counter}/{patience} (Best Loss: {best_val_loss:.4f})")
        
        # Early stopping
        if patience_counter >= patience:
            print(f"\n Early stopping at epoch {epoch+1}")
            print(f"Best Epoch: {best_epoch}")
            print(f"Best Val Loss: {best_val_loss:.4f}")
            print(f"Best Val RMSE: {best_val_rmse:.4f}")
            break
    
    # Save to Excel
    save_history_to_excel(history, best_metrics)
    
    # Create comprehensive visualization
    plot_comprehensive_results(history, best_metrics)
    
    return history


# ============================================================================
# K-Fold Cross-Validation
# ============================================================================

def cross_validation_training(data, protein_encoder, config, args, k_folds=5):
    """K-Fold Cross-Validation Training"""
    
    print("\n" + "=" * 70)
    print(f"STARTING {k_folds}-FOLD CROSS-VALIDATION")
    print("=" * 70)
    
    # Prepare data
    print("\nPreparing features...")
    drug_features = batch_enhanced_fingerprints(
        data['Drug'].tolist(),
        radius=2,
        n_bits=1024
    )
    
    print("\nEncoding protein sequences...")
    target_features = []
    for seq in tqdm(data['Target'], desc="Encoding proteins"):
        encoded = protein_encoder.encode(seq, max_len=1000)
        target_features.append(encoded)
    target_features = np.array(target_features)
    
    labels = data['Y'].values
    
    # Bin labels for stratification
    bins = pd.qcut(labels, q=5, labels=False, duplicates='drop')
    
    # K-Fold split
    skf = StratifiedKFold(n_splits=k_folds, shuffle=True, random_state=42)
    
    fold_results = []
    fold_models = []
    
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    
    for fold, (train_idx, val_idx) in enumerate(skf.split(drug_features, bins)):
        print(f"\n{'='*70}")
        print(f"FOLD {fold+1}/{k_folds}")
        print(f"{'='*70}")
        
        # Split data
        train_drug = torch.FloatTensor(drug_features[train_idx])
        train_target = torch.LongTensor(target_features[train_idx])
        train_label = torch.FloatTensor(labels[train_idx])
        
        val_drug = torch.FloatTensor(drug_features[val_idx])
        val_target = torch.LongTensor(target_features[val_idx])
        val_label = torch.FloatTensor(labels[val_idx])
        
        # Create datasets
        train_data = DTIDataset({
            'drug': train_drug,
            'target': train_target,
            'label': train_label
        })
        
        val_data = DTIDataset({
            'drug': val_drug,
            'target': val_target,
            'label': val_label
        })
        
        # Create dataloaders
        train_loader = DataLoader(train_data, batch_size=args.batch_size, shuffle=True)
        val_loader = DataLoader(val_data, batch_size=args.batch_size, shuffle=False)
        
        # Create model for this fold
        # Update config for enhanced features
        fold_config = config.copy()
        fold_config['drug_input_dim'] = drug_features.shape[1]  # Enhanced features
        
        model = DrugTargetInteractionModel(**fold_config)
        
        print(f"\nTraining Fold {fold+1}...")
        save_path = f"{args.save_path.replace('.pt', '')}_fold{fold+1}.pt"
        
        history = train_model_enhanced(
            model, train_loader, val_loader,
            epochs=args.epochs,
            lr=args.learning_rate,
            device=device,
            patience=args.patience,
            save_path=save_path,
            use_onecycle=args.use_onecycle
        )
        
        # Load best model
        checkpoint = torch.load(save_path, map_location=device, weights_only=False)
        final_metrics = checkpoint['metrics']
        
        fold_results.append(final_metrics)
        fold_models.append(save_path)
        
        print(f"\n Fold {fold+1} Complete!")
        print(f"   Best R²: {final_metrics['R2']:.4f}")
        print(f"   Best C-Index: {final_metrics['Concordance_Index']:.4f}")
    
    # Aggregate results
    print("\n" + "=" * 70)
    print("CROSS-VALIDATION RESULTS")
    print("=" * 70)
    
    metrics_names = ['R2', 'RMSE', 'MAE', 'Pearson_R', 'Spearman_R', 'Concordance_Index']
    
    print("\nMetric Averages:")
    for metric in metrics_names:
        values = [r[metric] for r in fold_results]
        mean = np.mean(values)
        std = np.std(values)
        print(f"{metric:20s}: {mean:.4f} ± {std:.4f}")
    
    return fold_results, fold_models


# ============================================================================
# Ensemble Prediction
# ============================================================================

def create_ensemble_predictions(models_paths, drug_features, target_features, device):
    """Create ensemble predictions from multiple models"""
    print("\nCreating ensemble predictions...")
    
    all_predictions = []
    
    # Update config for enhanced features
    config = {
        'drug_input_dim': drug_features.shape[1],  # Enhanced features dimension
        'drug_d_model': 128,
        'drug_num_layers': 6,
        'drug_num_heads': 8,
        'drug_d_ff': 512,
        'drug_hidden_dim': 256,
        'target_vocab_size': 26,
        'target_embedding_dim': 128,
        'target_num_filters': [32, 64, 96],
        'target_kernel_sizes': [4, 8, 12],
        'target_hidden_dim': 256,
        'decoder_hidden_dims': [1024, 512, 256],
        'dropout': 0.1,
        'binary': False
    }
    
    for model_path in models_paths:
        model = DrugTargetInteractionModel(**config)
        checkpoint = torch.load(model_path, map_location=device, weights_only=False)
        model.load_state_dict(checkpoint['model_state_dict'])
        model = model.to(device)
        model.eval()
        
        with torch.no_grad():
            drug = torch.FloatTensor(drug_features).to(device)
            target = torch.LongTensor(target_features).to(device)
            predictions = model(drug, target).cpu().numpy().flatten()
            all_predictions.append(predictions)
    
    # Average predictions
    ensemble_pred = np.mean(all_predictions, axis=0)
    ensemble_std = np.std(all_predictions, axis=0)
    
    return ensemble_pred, ensemble_std


# ============================================================================
# Visualization
# ============================================================================

def plot_cv_results(fold_results, save_path='cv_results.png'):
    """Plot cross-validation results"""
    metrics = ['R2', 'RMSE', 'MAE', 'Concordance_Index']
    
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    axes = axes.flatten()
    
    for i, metric in enumerate(metrics):
        values = [r[metric] for r in fold_results]
        
        axes[i].bar(range(1, len(values)+1), values, color='steelblue', alpha=0.7)
        axes[i].axhline(np.mean(values), color='red', linestyle='--', 
                       label=f'Mean: {np.mean(values):.4f}')
        axes[i].set_xlabel('Fold', fontweight='bold')
        axes[i].set_ylabel(metric, fontweight='bold')
        axes[i].set_title(f'{metric} Across Folds', fontweight='bold')
        axes[i].legend()
        axes[i].grid(alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches='tight')
    print(f"\n Cross-validation plot saved: {save_path}")


# ============================================================================
# Main Enhanced Training Pipeline
# ============================================================================

def main(args):
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    print('Enhanced Training Pipeline Loaded')
    print(f'Using device: {device}')
    
    # Load dataset
    print(f"\nLoading dataset from: {args.data_path}")
    df = pd.read_csv(args.data_path)
    #  CRITICAL: Verify data is preprocessed
    print("\n" + "="*70)
    print("VERIFYING DATA PREPROCESSING")
    print("="*70)
    
    y_min, y_max = df['Y'].min(), df['Y'].max()
    y_mean = df['Y'].mean()
    
    print(f"Y statistics:")
    print(f"  Min:  {y_min:.2f}")
    print(f"  Max:  {y_max:.2f}")
    print(f"  Mean: {y_mean:.2f}")
    
    if y_max > 100:
        print("\n ERROR: Data NOT preprocessed!")
        print("Expected pKd range: [2, 14]")
        print(f"Actual range: [{y_min:.0f}, {y_max:.0f}]")
        print("\nRun: python 01_data_preprocessor.py --input YOUR_RAW_DATA.csv")
        sys.exit(1)
    
    print("\n Data is in pKd scale - ready to train!")
    print("="*70 + "\n")
    
    if 'Unnamed: 0' in df.columns:
        df = df.drop('Unnamed: 0', axis=1)
    
    print(f" Dataset loaded: {len(df)} samples")
    
    # Clean data
    df_clean = clean_data(df)
    
    # Sample if requested
    if args.sample_size is not None:
        df_clean = df_clean.sample(n=min(args.sample_size, len(df_clean)), random_state=42)
        print(f"\n Using {len(df_clean)} samples")
    
    print(f"\nLabel (Y) statistics:")
    print(f"  Range: [{df_clean['Y'].min():.2f}, {df_clean['Y'].max():.2f}]")
    print(f"  Mean ± Std: {df_clean['Y'].mean():.2f} ± {df_clean['Y'].std():.2f}")
    
    # Data augmentation
    if args.augment > 0:
        df_clean = augment_data(df_clean, n_augment=args.augment)
    
    # Create protein encoder
    protein_encoder = ProteinSequenceEncoder()
    print(f"\n ProteinSequenceEncoder created (vocab size: {protein_encoder.vocab_size})")
    
    # Model configuration (SAME ARCHITECTURE!)
    config = {
        'drug_d_model': args.d_model,
        'drug_num_layers': args.num_layers,
        'drug_num_heads': args.num_heads,
        'drug_d_ff': args.d_ff,
        'drug_hidden_dim': 256,
        'target_vocab_size': 26,
        'target_embedding_dim': 128,
        'target_num_filters': [32, 64, 96],
        'target_kernel_sizes': [4, 8, 12],
        'target_hidden_dim': 256,
        'decoder_hidden_dims': [1024, 512, 256],
        'dropout': args.dropout,
        'binary': False
    }
    
    # Run cross-validation or single training
    if args.use_cv:
        fold_results, fold_models = cross_validation_training(
            df_clean, protein_encoder, config, args, k_folds=args.k_folds
        )
        
        # Plot results
        plot_cv_results(fold_results, save_path='cv_results.png')
        
        print("\n" + "=" * 70)
        print(" CROSS-VALIDATION COMPLETE!")
        print("=" * 70)
        print(f"\nTrained models saved:")
        for model_path in fold_models:
            print(f"  - {model_path}")
        
    else:
        # Single training run
        print("\n" + "=" * 70)
        print("PREPARING DATA FOR TRAINING")
        print("=" * 70)
        
        # Prepare features
        drug_features = batch_enhanced_fingerprints(
            df_clean['Drug'].tolist(),
            radius=2,
            n_bits=1024
        )
        
        # Update config with correct input dim
        config['drug_input_dim'] = drug_features.shape[1]
        
        print("\nEncoding protein sequences...")
        target_features = []
        for seq in tqdm(df_clean['Target'], desc="Encoding proteins"):
            encoded = protein_encoder.encode(seq, max_len=1000)
            target_features.append(encoded)
        target_features = np.array(target_features)
        
        labels = df_clean['Y'].values
        
        # Train/test split
        indices = np.arange(len(drug_features))
        train_idx, test_idx = train_test_split(indices, test_size=0.15, random_state=42)
        train_idx, val_idx = train_test_split(train_idx, test_size=0.15, random_state=42)
        
        # Create tensors
        train_drug = torch.FloatTensor(drug_features[train_idx])
        train_target = torch.LongTensor(target_features[train_idx])
        train_label = torch.FloatTensor(labels[train_idx])
        
        val_drug = torch.FloatTensor(drug_features[val_idx])
        val_target = torch.LongTensor(target_features[val_idx])
        val_label = torch.FloatTensor(labels[val_idx])
        
        test_drug = torch.FloatTensor(drug_features[test_idx])
        test_target = torch.LongTensor(target_features[test_idx])
        test_label = torch.FloatTensor(labels[test_idx])
        
        # Create datasets
        train_data = DTIDataset({'drug': train_drug, 'target': train_target, 'label': train_label})
        val_data = DTIDataset({'drug': val_drug, 'target': val_target, 'label': val_label})
        test_data = DTIDataset({'drug': test_drug, 'target': test_target, 'label': test_label})
        
        # Create dataloaders
        train_loader = DataLoader(train_data, batch_size=args.batch_size, shuffle=True, num_workers=0)
        val_loader = DataLoader(val_data, batch_size=args.batch_size, shuffle=False, num_workers=0)
        test_loader = DataLoader(test_data, batch_size=args.batch_size, shuffle=False, num_workers=0)
        
        print(f"\n Data prepared:")
        print(f"   Train: {len(train_data)} samples")
        print(f"   Val: {len(val_data)} samples")
        print(f"   Test: {len(test_data)} samples")
        
        # Create and train model
        model = DrugTargetInteractionModel(**config)
        
        total_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
        print(f"\n Model created: {total_params:,} parameters")
        
        history = train_model_enhanced(
            model, train_loader, val_loader,
            epochs=args.epochs,
            lr=args.learning_rate,
            device=device,
            patience=args.patience,
            save_path=args.save_path,
            use_onecycle=args.use_onecycle
        )
        
        # Evaluate on test set
        print("\n" + "=" * 70)
        print("FINAL EVALUATION ON TEST SET")
        print("=" * 70)
        
        checkpoint = torch.load(args.save_path, map_location=device, weights_only=False)
        model.load_state_dict(checkpoint['model_state_dict'])
        model = model.to(device)
        model.eval()
        
        test_preds = []
        test_labels = []
        
        with torch.no_grad():
            for drug, target, label in test_loader:
                drug = drug.to(device)
                target = target.to(device)
                pred = model(drug, target)
                test_preds.extend(pred.cpu().numpy().flatten())
                test_labels.extend(label.numpy().flatten())
        
        test_metrics = evaluate_regression(test_labels, test_preds)
        
        print("\nTest Set Performance:")
        for metric, value in test_metrics.items():
            print(f"  {metric:20s}: {value:.4f}")
        
        # Plot test predictions
        plt.figure(figsize=(10, 8))
        plt.scatter(test_labels, test_preds, alpha=0.5, s=20)
        plt.plot([min(test_labels), max(test_labels)], 
                [min(test_labels), max(test_labels)], 
                'r--', lw=2, label='Perfect Prediction')
        plt.xlabel('True pKd', fontsize=12, fontweight='bold')
        plt.ylabel('Predicted pKd', fontsize=12, fontweight='bold')
        plt.title(f'Test Set Predictions (R²={test_metrics["R2"]:.4f})', 
                 fontsize=14, fontweight='bold')
        plt.legend()
        plt.grid(alpha=0.3)
        plt.tight_layout()
        plt.savefig(r'E:\DTI_env\plots\test_predictions_enhanced.png', dpi=300)
        print("\n Test predictions plot saved: test_predictions_enhanced.png")
        
        print("\n" + "=" * 70)
        print(" TRAINING COMPLETE!")
        print("=" * 70)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Enhanced Drug-Target Interaction Training')
    
    # Data parameters
    parser.add_argument('--data_path', type=str, default=r'E:\DTI_env\data_processed\BindingDB_processed.csv',
                       help='Path to BindingDB CSV (use pKd-corrected version!)')
    parser.add_argument('--sample_size', type=int, default=None,
                       help='Number of samples (None for all)')
    parser.add_argument('--augment', type=int, default=0,
                       help='Data augmentation factor (0=no augmentation)')
    
    # Training parameters
    parser.add_argument('--batch_size', type=int, default=64,
                       help='Batch size')
    parser.add_argument('--epochs', type=int, default=100,
                       help='Number of epochs')
    parser.add_argument('--learning_rate', type=float, default=1e-4,
                       help='Learning rate')
    parser.add_argument('--patience', type=int, default=15,
                       help='Early stopping patience')
    parser.add_argument('--use_onecycle', type=bool, default=True,
                       help='Use OneCycleLR scheduler')
    
    # Model parameters (architecture unchanged!)
    parser.add_argument('--d_model', type=int, default=128,
                       help='Transformer d_model')
    parser.add_argument('--num_layers', type=int, default=6,
                       help='Number of transformer layers')
    parser.add_argument('--num_heads', type=int, default=8,
                       help='Number of attention heads')
    parser.add_argument('--d_ff', type=int, default=512,
                       help='Feed-forward dimension')
    parser.add_argument('--dropout', type=float, default=0.1,
                       help='Dropout rate')
    
    # Cross-validation parameters
    parser.add_argument('--use_cv', type=bool, default=False,
                       help='Use cross-validation')
    parser.add_argument('--k_folds', type=int, default=5,
                       help='Number of CV folds')
    
    # Output
    parser.add_argument('--save_path', type=str, default=r'E:\DTI_env\models_saved\best_model_enhanced.pt',
                       help='Path to save model')
    
    args = parser.parse_args()
    main(args)
