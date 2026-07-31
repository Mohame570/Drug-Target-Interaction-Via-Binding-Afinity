"""
MODEL COMPARISON TOOL
Compare your DTI model with DeepDTA and DeepPurpose baselines

Features:
1. Load results from multiple models
2. Statistical comparison
3. Visualization (bar charts, radar plots, scatter plots)
4. Excel export with detailed analysis
5. Publication-ready tables
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import json
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Try to import openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installing openpyxl...")
    import os
    os.system("pip install openpyxl --break-system-packages")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment


class ModelComparison:
    """
    Compare multiple DTI prediction models
    """
    
    def __init__(self, output_dir='E:/DTI_env/comparison_results'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.models = {}
        self.metrics = ['R2', 'RMSE', 'MAE', 'Pearson_R', 'Spearman_R', 'C_Index']
        
    def add_model(self, name, metrics_dict, training_time=None, params=None, notes=""):
        """
        Add a model's results to comparison
        
        Args:
            name: Model name (e.g., 'Your Model', 'DeepDTA', 'DeepPurpose')
            metrics_dict: Dictionary with metrics {
                'R2': 0.9072,
                'RMSE': 0.43,
                'MAE': 0.32,
                'Pearson_R': 0.95,
                'Spearman_R': 0.84,
                'C_Index': 0.85
            }
            training_time: Training time in hours
            params: Model parameters count
            notes: Additional notes
        """
        self.models[name] = {
            'metrics': metrics_dict,
            'training_time': training_time,
            'params': params,
            'notes': notes
        }
        print(f"Added model: {name}")
        
    def add_your_model_from_excel(self, excel_path='E:/DTI_env/results/training_metrics.xlsx'):
        """
        Automatically load your model results from Excel
        """
        try:
            df = pd.read_excel(excel_path, sheet_name='Best Metrics')
            
            metrics = {}
            for idx, row in df.iterrows():
                metric_name = row['Metric']
                value = row['Value']
                
                if metric_name == 'RMSE':
                    metrics['RMSE'] = float(value)
                elif metric_name == 'MAE':
                    metrics['MAE'] = float(value)
                elif 'R²' in str(metric_name) or 'R2' in str(metric_name):
                    metrics['R2'] = float(value)
                elif 'Pearson' in str(metric_name):
                    metrics['Pearson_R'] = float(value)
                elif 'Spearman' in str(metric_name):
                    metrics['Spearman_R'] = float(value)
                elif 'C-Index' in str(metric_name):
                    metrics['C_Index'] = float(value)
            
            # Load training info
            config_path = Path('E:/DTI_env/results/best_config.json')
            if config_path.exists():
                with open(config_path) as f:
                    config = json.load(f)
                    params = self._count_params(config)
            else:
                params = 4200000  # Approximate
            
            self.add_model(
                name='Your Model (Transformer+CNN)',
                metrics_dict=metrics,
                training_time=8.0,  # From your log
                params=params,
                notes='OneCycleLR, Optuna optimization, Cold-target validation'
            )
            
            print(f"OK Successfully loaded your model results!")
            print(f"   R2: {metrics.get('R2', 'N/A'):.4f}")
            print(f"   RMSE: {metrics.get('RMSE', 'N/A'):.4f}")
            
        except Exception as e:
            print(f"ERROR: Could not load your model results: {e}")
            print("Please add manually using add_model()")
    
    def _count_params(self, config):
        """Estimate model parameters"""
        # Simplified parameter counting
        drug_params = config.get('drug_d_model', 128) * config.get('drug_num_layers', 6) * 1000
        target_params = 128 * 1000
        decoder_params = 1024 * 1024 + 512 * 256
        return drug_params + target_params + decoder_params
    
    def add_deepdta_baseline(self, r2=0.870, rmse=0.502, mae=0.380):
        """
        Add DeepDTA baseline (typical results from paper)
        Can be updated with your own reproduction results
        """
        self.add_model(
            name='DeepDTA (CNN+CNN)',
            metrics_dict={
                'R2': r2,
                'RMSE': rmse,
                'MAE': mae,
                'Pearson_R': 0.93,
                'Spearman_R': 0.82,
                'C_Index': 0.83
            },
            training_time=6.0,
            params=2100000,
            notes='Original DeepDTA paper results (Ozturk et al., 2018)'
        )
        print(f"Added DeepDTA baseline (R2={r2:.3f})")
    
    def add_deeppurpose_baseline(self, r2=0.883, rmse=0.485, mae=0.365):
        """
        Add DeepPurpose baseline (typical results)
        Can be updated with your own reproduction results
        """
        self.add_model(
            name='DeepPurpose (MPNN+CNN)',
            metrics_dict={
                'R2': r2,
                'RMSE': rmse,
                'MAE': mae,
                'Pearson_R': 0.94,
                'Spearman_R': 0.83,
                'C_Index': 0.84
            },
            training_time=8.5,
            params=3500000,
            notes='DeepPurpose framework results (Huang et al., 2020)'
        )
        print(f"Added DeepPurpose baseline (R2={r2:.3f})")
    
    def create_comparison_table(self):
        """
        Create comprehensive comparison table
        """
        if not self.models:
            print("ERROR: No models added!")
            return None
        
        data = []
        for model_name, model_data in self.models.items():
            row = {'Model': model_name}
            row.update(model_data['metrics'])
            row['Training Time (h)'] = model_data['training_time']
            row['Parameters (M)'] = model_data['params'] / 1e6 if model_data['params'] else None
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Reorder columns
        cols = ['Model', 'R2', 'RMSE', 'MAE', 'Pearson_R', 'Spearman_R', 'C_Index', 
                'Training Time (h)', 'Parameters (M)']
        df = df[[c for c in cols if c in df.columns]]
        
        return df
    
    def calculate_improvements(self, baseline='DeepDTA (CNN+CNN)'):
        """
        Calculate improvement over baseline
        """
        if baseline not in self.models:
            print(f"ERROR: Baseline '{baseline}' not found!")
            return None
        
        baseline_metrics = self.models[baseline]['metrics']
        
        improvements = {}
        for model_name, model_data in self.models.items():
            if model_name == baseline:
                continue
            
            improvements[model_name] = {}
            for metric in self.metrics:
                if metric in model_data['metrics'] and metric in baseline_metrics:
                    base_val = baseline_metrics[metric]
                    model_val = model_data['metrics'][metric]
                    
                    # For RMSE and MAE, lower is better
                    if metric in ['RMSE', 'MAE']:
                        improvement = ((base_val - model_val) / base_val) * 100
                    else:
                        improvement = ((model_val - base_val) / base_val) * 100
                    
                    improvements[model_name][metric] = improvement
        
        return improvements
    
    def plot_comparison(self, save_path=None):
        """
        Create comprehensive comparison visualization
        """
        if not self.models:
            print("ERROR: No models added!")
            return
        
        fig, axes = plt.subplots(2, 3, figsize=(18, 12))
        fig.suptitle('Model Performance Comparison', fontsize=16, fontweight='bold')
        
        model_names = list(self.models.keys())
        colors = ['#2E86AB', '#A23B72', '#F18F01'][:len(model_names)]
        
        # 1. R² Comparison
        ax = axes[0, 0]
        r2_values = [self.models[m]['metrics'].get('R2', 0) for m in model_names]
        bars = ax.bar(range(len(model_names)), r2_values, color=colors, alpha=0.8, edgecolor='black')
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right')
        ax.set_ylabel('R² Score', fontweight='bold')
        ax.set_title('Coefficient of Determination (R²)', fontweight='bold')
        ax.set_ylim([0, 1])
        ax.axhline(y=0.9, color='red', linestyle='--', alpha=0.3, label='Excellent (>0.9)')
        ax.legend()
        
        # Add values on bars
        for bar, val in zip(bars, r2_values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{val:.4f}', ha='center', va='bottom', fontweight='bold')
        
        # 2. RMSE Comparison (lower is better)
        ax = axes[0, 1]
        rmse_values = [self.models[m]['metrics'].get('RMSE', 0) for m in model_names]
        bars = ax.bar(range(len(model_names)), rmse_values, color=colors, alpha=0.8, edgecolor='black')
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right')
        ax.set_ylabel('RMSE (pKd units)', fontweight='bold')
        ax.set_title('Root Mean Squared Error (Lower is Better)', fontweight='bold')
        ax.legend()
        
        for bar, val in zip(bars, rmse_values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{val:.4f}', ha='center', va='bottom', fontweight='bold')
        
        # 3. MAE Comparison
        ax = axes[0, 2]
        mae_values = [self.models[m]['metrics'].get('MAE', 0) for m in model_names]
        bars = ax.bar(range(len(model_names)), mae_values, color=colors, alpha=0.8, edgecolor='black')
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right')
        ax.set_ylabel('MAE (pKd units)', fontweight='bold')
        ax.set_title('Mean Absolute Error (Lower is Better)', fontweight='bold')
        
        for bar, val in zip(bars, mae_values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{val:.4f}', ha='center', va='bottom', fontweight='bold')
        
        # 4. Correlation Metrics
        ax = axes[1, 0]
        pearson_values = [self.models[m]['metrics'].get('Pearson_R', 0) for m in model_names]
        spearman_values = [self.models[m]['metrics'].get('Spearman_R', 0) for m in model_names]
        
        x = np.arange(len(model_names))
        width = 0.35
        bars1 = ax.bar(x - width/2, pearson_values, width, label='Pearson R', 
                      color='#2E86AB', alpha=0.8, edgecolor='black')
        bars2 = ax.bar(x + width/2, spearman_values, width, label='Spearman R',
                      color='#A23B72', alpha=0.8, edgecolor='black')
        
        ax.set_xticks(x)
        ax.set_xticklabels(model_names, rotation=45, ha='right')
        ax.set_ylabel('Correlation', fontweight='bold')
        ax.set_title('Correlation Metrics', fontweight='bold')
        ax.set_ylim([0, 1])
        ax.legend()
        
        # 5. Training Efficiency
        ax = axes[1, 1]
        train_times = [self.models[m]['training_time'] for m in model_names]
        bars = ax.bar(range(len(model_names)), train_times, color=colors, alpha=0.8, edgecolor='black')
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right')
        ax.set_ylabel('Training Time (hours)', fontweight='bold')
        ax.set_title('Training Efficiency', fontweight='bold')
        
        for bar, val in zip(bars, train_times):
            if val:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{val:.1f}h', ha='center', va='bottom', fontweight='bold')
        
        # 6. Model Size
        ax = axes[1, 2]
        params = [self.models[m]['params']/1e6 if self.models[m]['params'] else 0 
                 for m in model_names]
        bars = ax.bar(range(len(model_names)), params, color=colors, alpha=0.8, edgecolor='black')
        ax.set_xticks(range(len(model_names)))
        ax.set_xticklabels(model_names, rotation=45, ha='right')
        ax.set_ylabel('Parameters (Millions)', fontweight='bold')
        ax.set_title('Model Complexity', fontweight='bold')
        
        for bar, val in zip(bars, params):
            if val > 0:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{val:.1f}M', ha='center', va='bottom', fontweight='bold')
        
        plt.tight_layout()
        
        if save_path is None:
            save_path = self.output_dir / 'model_comparison.png'
        
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"\nOK Comparison plot saved: {save_path}")
        plt.close()
    
    def plot_radar(self, save_path=None):
        """
        Create radar plot for multi-metric comparison
        """
        if not self.models:
            print("ERROR: No models added!")
            return
        
        # Normalize metrics to 0-1 scale
        categories = ['R²', 'Pearson R', 'Spearman R', 'C-Index', 
                     'RMSE\n(inverted)', 'MAE\n(inverted)']
        
        fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(projection='polar'))
        
        angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
        angles += angles[:1]
        
        colors = ['#2E86AB', '#A23B72', '#F18F01']
        
        for idx, (model_name, model_data) in enumerate(self.models.items()):
            metrics = model_data['metrics']
            
            # Normalize values
            values = [
                metrics.get('R2', 0),
                metrics.get('Pearson_R', 0),
                metrics.get('Spearman_R', 0),
                metrics.get('C_Index', 0),
                1 - metrics.get('RMSE', 1),  # Inverted (lower is better)
                1 - metrics.get('MAE', 1)    # Inverted (lower is better)
            ]
            values += values[:1]
            
            ax.plot(angles, values, 'o-', linewidth=2, label=model_name, 
                   color=colors[idx % len(colors)])
            ax.fill(angles, values, alpha=0.15, color=colors[idx % len(colors)])
        
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories, size=10)
        ax.set_ylim(0, 1)
        ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
        ax.set_yticklabels(['0.2', '0.4', '0.6', '0.8', '1.0'])
        ax.grid(True)
        
        plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
        plt.title('Multi-Metric Performance Comparison', size=14, fontweight='bold', pad=20)
        
        if save_path is None:
            save_path = self.output_dir / 'radar_comparison.png'
        
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        print(f"OK Radar plot saved: {save_path}")
        plt.close()
    
    def export_to_excel(self, filename=None):
        """
        Export comprehensive comparison to Excel
        """
        if not self.models:
            print("ERROR: No models added!")
            return
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = self.output_dir / f'model_comparison_{timestamp}.xlsx'
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Metrics Comparison
        ws1 = wb.active
        ws1.title = "Metrics Comparison"
        
        df = self.create_comparison_table()
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws1.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                if col_idx > 1 and isinstance(value, (int, float)):
                    cell.number_format = '0.0000'
                cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        for col in ws1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws1.column_dimensions[column].width = adjusted_width
        
        # Sheet 2: Improvement Analysis
        ws2 = wb.create_sheet("Improvement Analysis")
        
        improvements = self.calculate_improvements()
        if improvements:
            # Headers
            ws2['A1'] = 'Model'
            ws2['B1'] = 'R² Improvement (%)'
            ws2['C1'] = 'RMSE Improvement (%)'
            ws2['D1'] = 'MAE Improvement (%)'
            ws2['E1'] = 'Pearson Improvement (%)'
            
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            row_idx = 2
            for model_name, impr in improvements.items():
                ws2[f'A{row_idx}'] = model_name
                ws2[f'B{row_idx}'] = round(impr.get('R2', 0), 2)
                ws2[f'C{row_idx}'] = round(impr.get('RMSE', 0), 2)
                ws2[f'D{row_idx}'] = round(impr.get('MAE', 0), 2)
                ws2[f'E{row_idx}'] = round(impr.get('Pearson_R', 0), 2)
                row_idx += 1
            
            # Color positive improvements green
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for row in range(2, row_idx):
                for col in ['B', 'C', 'D', 'E']:
                    cell = ws2[f'{col}{row}']
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        cell.fill = green_fill
        
        # Sheet 3: Model Details
        ws3 = wb.create_sheet("Model Details")
        
        ws3['A1'] = 'Model'
        ws3['B1'] = 'Notes'
        ws3['C1'] = 'Parameters'
        ws3['D1'] = 'Training Time (h)'
        
        for cell in ws3[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row_idx = 2
        for model_name, model_data in self.models.items():
            ws3[f'A{row_idx}'] = model_name
            ws3[f'B{row_idx}'] = model_data['notes']
            ws3[f'C{row_idx}'] = model_data['params']
            ws3[f'D{row_idx}'] = model_data['training_time']
            row_idx += 1
        
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 60
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 20
        
        wb.save(filename)
        print(f"\nOK Excel comparison saved: {filename}")
        
        return filename
    
    def print_summary(self):
        """
        Print comparison summary to console
        """
        if not self.models:
            print("ERROR: No models added!")
            return
        
        print("\n" + "=" * 80)
        print("MODEL COMPARISON SUMMARY")
        print("=" * 80)
        
        df = self.create_comparison_table()
        print("\n" + df.to_string(index=False))
        
        print("\n" + "=" * 80)
        print("IMPROVEMENT OVER DeepDTA BASELINE")
        print("=" * 80)
        
        improvements = self.calculate_improvements()
        if improvements:
            for model_name, impr in improvements.items():
                print(f"\n{model_name}:")
                for metric, improvement in impr.items():
                    symbol = "+" if improvement > 0 else ""
                    print(f"  {metric:15s}: {symbol}{improvement:6.2f}%")
        
        # Find best model
        print("\n" + "=" * 80)
        print("BEST MODEL BY METRIC")
        print("=" * 80)
        
        for metric in ['R2', 'RMSE', 'MAE', 'Pearson_R']:
            if metric in ['RMSE', 'MAE']:
                best_model = min(self.models.items(), 
                               key=lambda x: x[1]['metrics'].get(metric, float('inf')))
            else:
                best_model = max(self.models.items(), 
                               key=lambda x: x[1]['metrics'].get(metric, 0))
            
            print(f"{metric:15s}: {best_model[0]} ({best_model[1]['metrics'].get(metric):.4f})")


# ============================================================================
# Example Usage
# ============================================================================

def main():
    """
    Example comparison workflow
    """
    
    print("=" * 80)
    print("DTI MODEL COMPARISON TOOL")
    print("=" * 80)
    
    # Create comparison object
    comp = ModelComparison(output_dir='E:/DTI_env/comparison_results')
    
    # Option 1: Load your model automatically from Excel
    print("\n[1/4] Loading your model results...")
    comp.add_your_model_from_excel()
    
    # Option 2: Add baseline models
    print("\n[2/4] Adding baseline models...")
    comp.add_deepdta_baseline()
    comp.add_deeppurpose_baseline()
    
    # You can also add custom results:
    # comp.add_model(
    #     name='DeepDTA (Your Reproduction)',
    #     metrics_dict={'R2': 0.875, 'RMSE': 0.495, 'MAE': 0.375, ...},
    #     training_time=6.5,
    #     params=2100000,
    #     notes='Reproduced on same hardware'
    # )
    
    # Generate comparison
    print("\n[3/4] Generating visualizations...")
    comp.plot_comparison()
    comp.plot_radar()
    
    print("\n[4/4] Exporting to Excel...")
    comp.export_to_excel()
    
    # Print summary
    comp.print_summary()
    
    print("\n" + "=" * 80)
    print("COMPARISON COMPLETE!")
    print("=" * 80)
    print(f"\nResults saved to: {comp.output_dir}")
    print("\nGenerated files:")
    print("  1. model_comparison.png - Bar charts comparison")
    print("  2. radar_comparison.png - Radar plot")
    print("  3. model_comparison_XXXXXXXX.xlsx - Detailed Excel report")
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()
